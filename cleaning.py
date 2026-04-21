"""
cleaning.py — Generic, project-agnostic data-cleaning engine for Decipher
(Forsta) survey data.  All business rules are driven by a config dict
supplied at runtime.  No hardcoded variable names or thresholds.

Dependencies: requests, pandas, numpy, openpyxl
"""

import json
import re
import requests
import pandas as pd
import numpy as np
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook


# ──────────────────────────────────────────────────────────────────────
# Internal helpers
# ──────────────────────────────────────────────────────────────────────

def _safe_list(val):
    """Return val if it is a list, otherwise return [].
    Prevents 'NoneType is not iterable' when a config array is null."""
    return val if isinstance(val, list) else []


def _ensure_dict(config):
    """Return config as a dict.
    Guards against the config being passed as a JSON string."""
    if isinstance(config, str):
        return json.loads(config)
    if config is None:
        return {}
    return config


def clean_cast_column(series: pd.Series) -> pd.Series:
    """Cast a column to the tightest appropriate dtype.
    Strings that look numeric become int64 or float64; everything else
    stays as nullable string.  Sentinel strings are normalised to pd.NA.
    """
    series = series.replace(["<NA>", "nan", "NaN", "None", ""], pd.NA)
    numeric = pd.to_numeric(series, errors="coerce")
    non_numeric_mask = series.notna() & numeric.isna()

    if non_numeric_mask.any():
        return series.astype("string")

    if numeric.notna().any():
        is_all_whole = (numeric.dropna() == numeric.dropna().astype(int)).all()
        has_nulls = numeric.isna().any()
        if is_all_whole and not has_nulls:
            return numeric.astype("int64")
        else:
            return numeric.astype("float64")

    return series.astype("string")


def ensure_columns(df: pd.DataFrame, cols: list, fill_value=np.nan) -> pd.DataFrame:
    """Guarantee that every column in *cols* exists in *df*.
    Missing columns are created and filled with *fill_value*.
    Decipher API omits columns that have zero responses, so this guard
    is essential before any column-level operations.
    """
    for c in [col for col in cols if col not in df.columns]:
        df[c] = fill_value
    return df


def _resolve_grid_vars(grid: dict) -> list:
    """Resolve a grid's variable list.
    Uses explicit 'vars' list if provided and non-empty.
    Falls back to generating from 'prefix' + 'max_idx' if vars is null/[].
    Returns [] if neither is available.
    """
    vars_list = _safe_list(grid.get("vars"))
    if vars_list:
        return vars_list
    prefix  = grid.get("prefix") or ""
    max_idx = grid.get("max_idx")
    if prefix and max_idx:
        return [f"{prefix}{i}" for i in range(1, int(max_idx) + 1)]
    return []


# ──────────────────────────────────────────────────────────────────────
# Column ordering helpers
# ──────────────────────────────────────────────────────────────────────

def _datamap_sort_key(col_name: str, dm_index: dict, fallback: int) -> tuple:
    """Return a sort key (position, suffix) for *col_name* based on the
    column_order index from the config.

    For exact matches, returns (position, 0).
    For expanded grid columns (e.g. Parent_2r3c2 → parent Parent_2),
    finds the longest matching prefix and returns (parent_pos, suffix)
    so that child columns stay grouped under their parent in order.
    Unmatched columns sort to *fallback* (end).
    """
    if col_name in dm_index:
        return (dm_index[col_name], 0)

    # Try progressively shorter prefixes to find the parent variable
    best_pos = fallback
    best_len = 0
    for var, pos in dm_index.items():
        if col_name.startswith(var) and len(var) > best_len:
            best_pos = pos
            best_len = len(var)

    # Use the remaining suffix as secondary sort so children stay ordered
    suffix = col_name[best_len:] if best_len > 0 else col_name
    nums = re.findall(r"\d+", suffix)
    suffix_key = tuple(int(n) for n in nums) if nums else (0,)
    return (best_pos, *suffix_key)


# ──────────────────────────────────────────────────────────────────────
# Data pull
# ──────────────────────────────────────────────────────────────────────

def pull_data(
    survey_id: str,
    client_id: str,
    api_key: str,
    start_date: str = None,
) -> pd.DataFrame:
    """Fetch qualified (status == 3) responses from Decipher and return a
    type-cast DataFrame.  Optionally filter to responses on or after
    *start_date* (ISO-8601 string).
    """
    headers = {"x-apikey": api_key, "Accept": "application/json"}
    url = (
        f"https://sw2.decipherinc.com/api/v1/surveys/selfserve/"
        f"{client_id}/{survey_id}/data?format=json"
    )

    response = requests.get(url, headers=headers, timeout=120)

    if response.status_code == 401:
        raise PermissionError("Could not authenticate. Check the API key.")
    if response.status_code == 404:
        raise FileNotFoundError(f"Survey '{survey_id}' not found in Decipher.")
    response.raise_for_status()

    data = response.json()
    if not data:
        raise ValueError("API returned an empty dataset.")

    df = pd.json_normalize(data)
    df = df.apply(clean_cast_column)

    df = ensure_columns(df, ["status"])
    df = df[df["status"] == 3].copy()
    df = df.apply(clean_cast_column)

    if df.empty:
        raise ValueError("No qualified (status = 3) responses in this survey.")

    if start_date:
        df = ensure_columns(df, ["start_date"])
        df["start_date"] = pd.to_datetime(df["start_date"], errors="coerce")
        df = df[df["start_date"] >= pd.to_datetime(start_date)].copy()
        if df.empty:
            raise ValueError(
                f"No qualified responses found from {start_date} onwards."
            )

    df.reset_index(drop=True, inplace=True)
    return df


# ──────────────────────────────────────────────────────────────────────
# Flag engine
# ──────────────────────────────────────────────────────────────────────

def apply_standard_flags(df: pd.DataFrame, config: dict) -> pd.DataFrame:
    """Apply every configured flag family and return the augmented DataFrame."""
    config = _ensure_dict(config)

    # ── LOI / Speeding ────────────────────────────────────────────────
    loi_var      = config.get("loi_variable", "qtime")
    loi_pct      = config.get("loi_min_threshold_pct", 0.33)
    exclude_if_oe = config.get("loi_exclude_if_quality_oe", False)
    oe_cols      = [v["var"] for v in _safe_list(config.get("oe_variables"))]

    df = ensure_columns(df, [loi_var] + oe_cols)

    loi_numeric = pd.to_numeric(df[loi_var], errors="coerce")
    median_loi  = loi_numeric.median()

    df["flag_speeder"] = 0.0
    df.loc[loi_numeric < (median_loi * loi_pct), "flag_speeder"] = 1.0

    df["flag_lagger"] = 0.0
    df.loc[loi_numeric > (median_loi * 3), "flag_lagger"] = 0.5

    if exclude_if_oe and oe_cols:
        existing_oe = [c for c in oe_cols if c in df.columns]
        if existing_oe:
            has_oe = (
                df[existing_oe].notna().any(axis=1) &
                df[existing_oe].astype(str).replace("", pd.NA).notna().any(axis=1)
            )
            df.loc[(df["flag_speeder"] == 1.0) & has_oe, "flag_speeder"] = 0.5

    # ── Straightlining ────────────────────────────────────────────────
    sl_flag_cols = []

    for grid in _safe_list(config.get("grid_variables")):
        grid_vars = _resolve_grid_vars(grid)
        if not grid_vars:
            continue

        label     = grid.get("label", "unknown")
        min_items = int(grid.get("min_straightline_items") or 3)
        col_name  = f"flag_SL_{label}"
        sl_flag_cols.append(col_name)

        df = ensure_columns(df, grid_vars)

        def _check_sl(row, gv=grid_vars, mi=min_items):
            vals         = row[gv].dropna()
            vals_numeric = pd.to_numeric(vals, errors="coerce")
            use          = vals_numeric if vals_numeric.notna().all() else vals
            if len(use) < mi:
                return 0.0
            return 0.5 if use.nunique() == 1 else 0.0

        df[col_name] = df.apply(_check_sl, axis=1)

    df["SL_Count"] = (
        (df[sl_flag_cols] != 0).sum(axis=1) if sl_flag_cols else 0
    )

    # ── Inconsistency checks ─────────────────────────────────────────
    for chk in _safe_list(config.get("inconsistency_checks")):
        # Skip manual-only checks — they require human review
        if chk.get("manual_check", False):
            continue

        cid          = chk.get("id", "unknown")
        if_var       = chk.get("if_var", "")
        if_val       = chk.get("if_val")
        then_var     = chk.get("then_var", "")
        invalid_vals = _safe_list(chk.get("invalid_vals"))
        col_name     = f"flag_{cid}"

        if not if_var or not then_var or not invalid_vals:
            # Not enough info to evaluate — skip silently
            df[col_name] = 0.0
            continue

        df = ensure_columns(df, [if_var, then_var])

        def _check_inc(row, iv=if_var, ivl=if_val, tv=then_var, bad=invalid_vals):
            if pd.isna(row[iv]) or pd.isna(row[tv]):
                return 0.0
            try:
                row_iv = int(float(row[iv]))
                row_tv = int(float(row[tv]))
            except (ValueError, TypeError):
                row_iv = row[iv]
                row_tv = row[tv]
            if ivl is None:
                return 0.0
            if row_iv == ivl and row_tv in bad:
                return 1.0
            return 0.0

        df[col_name] = df.apply(_check_inc, axis=1)

    # ── Numeric variable review ───────────────────────────────────────
    for entry in _safe_list(config.get("numeric_variables")):
        var      = entry.get("var", "")
        rmin     = entry.get("review_min")
        rmax     = entry.get("review_max")
        col_name = f"flag_{var}"

        if not var:
            continue

        df = ensure_columns(df, [var])
        numeric_vals = pd.to_numeric(df[var], errors="coerce")

        conditions = []
        if rmin is not None:
            conditions.append(numeric_vals < rmin)
        if rmax is not None:
            conditions.append(numeric_vals > rmax)

        if conditions:
            # Use fillna(False) so pd.NA comparisons resolve to False
            outside = conditions[0].fillna(False)
            for c in conditions[1:]:
                outside = outside | c.fillna(False)
            df[col_name] = np.where(numeric_vals.isna(), 0.0,
                                    np.where(outside, 0.5, 0.0))
        else:
            df[col_name] = 0.0

    # ── MaxDiff timer check ───────────────────────────────────────────
    df["flag_maxdiff_speed"] = 0.0
    for sec in _safe_list(config.get("maxdiff_sections")):
        timer_var = sec.get("timer_var", "")
        min_sec   = sec.get("min_seconds")
        if not timer_var or min_sec is None:
            continue
        df = ensure_columns(df, [timer_var])
        timer_vals = pd.to_numeric(df[timer_var], errors="coerce")
        df.loc[timer_vals < min_sec, "flag_maxdiff_speed"] = 1.0

    return df


# ──────────────────────────────────────────────────────────────────────
# Summary columns
# ──────────────────────────────────────────────────────────────────────

def build_summary_columns(df: pd.DataFrame, config: dict = None) -> pd.DataFrame:
    """Add total_flags, flag_remove, flag_review, affected_questions,
    reason, and action columns.

    If *config* is supplied, flag labels in the affected_questions column
    use human-readable names from the config instead of raw column names.
    """
    config       = _ensure_dict(config) if config else {}
    label_map    = _build_rename_map(config)   # flag_incon_1 → "label text"
    flag_cols    = [c for c in df.columns if c.startswith("flag_")]
    sl_flag_cols = [c for c in df.columns if c.startswith("flag_SL_")]

    df["total_flags"] = df[flag_cols].sum(axis=1) if flag_cols else 0.0

    if sl_flag_cols:
        df["SL_Count"] = (df[sl_flag_cols] != 0).sum(axis=1)
    elif "SL_Count" not in df.columns:
        df["SL_Count"] = 0

    df["flag_remove"] = 0
    if "flag_speeder" in df.columns:
        df.loc[df["flag_speeder"] == 1.0, "flag_remove"] = 1
    if "flag_maxdiff_speed" in df.columns:
        df.loc[df["flag_maxdiff_speed"] == 1.0, "flag_remove"] = 1

    df["flag_review"] = np.where(
        (df["total_flags"] > 0) & (df["flag_remove"] == 0), 1, 0
    )

    def _build_affected(row):
        parts        = []
        detail_parts = []

        if row.get("flag_speeder", 0) > 0:
            parts.append("Speeder")
            detail_parts.append("Speeder")

        if row.get("flag_lagger", 0) > 0:
            parts.append("Lagger")
            detail_parts.append("Lagger")

        inc_only = []
        num_only = []
        for f in flag_cols:
            if f in ("flag_speeder", "flag_lagger", "flag_maxdiff_speed",
                     "flag_remove", "flag_review"):
                continue
            if f.startswith("flag_SL_"):
                continue
            val = row.get(f, 0)
            if val <= 0:
                continue
            # Use the human-readable label if available, else raw name
            display = label_map.get(f, f)
            if val == 0.5:
                num_only.append(display)
            else:
                inc_only.append(display)

        if inc_only:
            names = ", ".join(inc_only)
            parts.append(f"Inconsistent/Illogical Response ({names})")
            detail_parts.append("Inconsistent/Illogical Response")

        if num_only:
            names = ", ".join(num_only)
            parts.append(f"Numeric Out-of-Range ({names})")
            detail_parts.append("Numeric Out-of-Range")

        fired_sl = [
            label_map.get(c, c.replace("flag_SL_", ""))
            for c in sl_flag_cols
            if row.get(c, 0) > 0
        ]
        if fired_sl:
            parts.append(f"Straightliner ({', '.join(fired_sl)})")
            detail_parts.append("Straightliner")

        if row.get("flag_maxdiff_speed", 0) > 0:
            parts.append("MaxDiff Speeder")
            detail_parts.append("MaxDiff Speeder")

        return ("; ".join(parts), "; ".join(detail_parts))

    affected = df.apply(_build_affected, axis=1, result_type="expand")
    df["affected_questions"] = affected[0].fillna("")
    df["reason"]             = affected[1].fillna("")
    df["action"]             = np.where(df["flag_remove"] == 1, "Remove", "")

    return df


# ──────────────────────────────────────────────────────────────────────
# OE extraction
# ──────────────────────────────────────────────────────────────────────

def build_oe_dataframe(df: pd.DataFrame, config: dict) -> pd.DataFrame:
    """Return a separate DataFrame with uuid + all configured OE columns.
    This feeds the OE Review sheet which the OE Review artifact reads.
    """
    config  = _ensure_dict(config)
    oe_vars = [v["var"] for v in _safe_list(config.get("oe_variables"))]
    oe_labels = {
        v["var"]: v.get("label", v["var"])
        for v in _safe_list(config.get("oe_variables"))
    }
    keep = ["uuid"] + oe_vars
    df   = ensure_columns(df, keep)
    return df[keep].copy(), oe_labels


# ──────────────────────────────────────────────────────────────────────
# IP risk check (Scamalytics)
# ──────────────────────────────────────────────────────────────────────

_SCAM_DELAY = (0.6, 1.2)
_SCAM_RETRIES = 3
_SCAM_BACKOFF = 2.0

import time
import random


def _scam_fetch(ip: str, username: str, key: str) -> dict:
    """Call Scamalytics API for a single IP with retry + backoff."""
    import time, random
    base_url = f"https://api11.scamalytics.com/v3/{username}/"
    for attempt in range(_SCAM_RETRIES):
        try:
            r = requests.get(
                base_url, params={"key": key, "ip": ip},
                headers={"User-Agent": "survey-cleaning/1.0", "Accept": "application/json"},
                timeout=20,
            )
            if r.status_code == 200:
                try:
                    return {"ok": True, "json": r.json()}
                except Exception:
                    return {"ok": True, "json": {"_raw": r.text}}
            if r.status_code == 429:
                time.sleep(_SCAM_BACKOFF ** attempt + random.uniform(0, 1))
                continue
            if r.status_code in (401, 403):
                return {"ok": False, "error": f"Auth error {r.status_code}"}
            time.sleep(_SCAM_BACKOFF ** attempt + random.uniform(0, 1))
        except requests.RequestException as e:
            time.sleep(_SCAM_BACKOFF ** attempt + random.uniform(0, 1))
    return {"ok": False, "error": "Max retries"}


def _nested(d, keys, default=None):
    cur = d
    for k in keys:
        if not isinstance(cur, dict):
            return default
        cur = cur.get(k, default)
    return cur


def _to_bool(v) -> bool:
    if v is None or v is False:
        return False
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return bool(v)
    s = str(v).strip().lower()
    return s in ("1", "true", "t", "yes", "y", "on") or any(
        k in s for k in ("vpn", "tor", "datacenter", "proxy", "hosting")
    )


def _curate_ip(ip: str, p: dict) -> dict:
    """Extract curated fields from a Scamalytics API response."""
    b = lambda *keys: _to_bool(_nested(p, list(keys)))

    sc = "scamalytics"
    px = "scamalytics_proxy"
    ext = "external_datasources"

    score = _nested(p, [sc, "scamalytics_score"])
    risk = _nested(p, [sc, "scamalytics_risk"], "")

    sc_vpn = b(sc, px, "is_vpn")
    sc_dc = b(sc, px, "is_datacenter")
    x4_vpn = b(ext, "x4bnet", "is_vpn")
    x4_dc = b(ext, "x4bnet", "is_datacenter")
    x4_tor = b(ext, "x4bnet", "is_tor")

    ip2_type = str(_nested(p, [ext, "ip2proxy_lite", "proxy_type"], "")).lower()
    ip2_black = b(ext, "ip2proxy_lite", "ip_blacklisted")
    ipsum_black = b(ext, "ipsum", "ip_blacklisted")
    spamhaus_black = b(ext, "spamhaus_drop", "ip_blacklisted")
    firehol_proxy = b(ext, "firehol", "is_proxy")
    firehol_30 = b(ext, "firehol", "ip_blacklisted_30")
    firehol_1d = b(ext, "firehol", "ip_blacklisted_1day")

    return {
        "ip": ip,
        "risk": risk,
        "score": score,
        "isp_risk": _nested(p, [sc, "scamalytics_isp_risk"], ""),
        "is_vpn": sc_vpn or x4_vpn or ("vpn" in ip2_type),
        "is_datacenter": sc_dc or x4_dc or any(k in ip2_type for k in ("datacenter", "hosting")),
        "is_tor": x4_tor or ("tor" in ip2_type),
        "is_proxy": any([firehol_proxy, ip2_black, ipsum_black, spamhaus_black, sc_vpn, sc_dc]),
        "is_blacklisted": any([ip2_black, ipsum_black, spamhaus_black, firehol_1d, firehol_30]),
        "is_bot": any([
            _to_bool(_nested(p, [sc, "bot_status"])),
            b(ext, "x4bnet", "is_bot_operamini"),
            b(ext, "x4bnet", "is_bot_semrush"),
            b(ext, "google", "is_googlebot"),
        ]),
        "is_icloud_relay": b(sc, px, "is_apple_icloud_private_relay"),
    }


_EMPTY_IP_ROW = {
    "ip": "", "risk": "", "score": None, "isp_risk": "",
    "is_vpn": False, "is_datacenter": False, "is_tor": False,
    "is_proxy": False, "is_blacklisted": False, "is_bot": False,
    "is_icloud_relay": False, "error": "",
}


# ── GeoLite2 geolocation ─────────────────────────────────────────────

_GEOIP_CACHE_PATH = Path("/tmp/GeoLite2-City.mmdb")


def ensure_geoip_db(license_key: str = None, local_path: str = None) -> str:
    """Return a valid path to a GeoLite2-City.mmdb file.

    Resolution order:
      1. *local_path* if it exists on disk (e.g. committed to repo)
      2. Cached download at /tmp/GeoLite2-City.mmdb
      3. Fresh download from MaxMind using *license_key*

    Returns the path string, or "" if no database is available.
    """
    # 1. Local file provided and exists
    if local_path and Path(local_path).is_file():
        return str(local_path)

    # 2. Already cached from a previous run
    if _GEOIP_CACHE_PATH.is_file():
        return str(_GEOIP_CACHE_PATH)

    # 3. Download from MaxMind
    if not license_key:
        return ""

    import tarfile
    import io

    url = (
        "https://download.maxmind.com/app/geoip_download"
        f"?edition_id=GeoLite2-City&license_key={license_key}&suffix=tar.gz"
    )
    try:
        r = requests.get(url, timeout=120)
        r.raise_for_status()
        with tarfile.open(fileobj=io.BytesIO(r.content), mode="r:gz") as tar:
            for member in tar.getmembers():
                if member.name.endswith(".mmdb"):
                    f = tar.extractfile(member)
                    if f:
                        _GEOIP_CACHE_PATH.write_bytes(f.read())
                        return str(_GEOIP_CACHE_PATH)
    except Exception:
        pass
    return ""


def _geoip_enrich(ip_df: pd.DataFrame, db_path: str) -> pd.DataFrame:
    """Add country, state, city columns to ip_df using a local GeoLite2-City
    database.  Requires the ``geoip2`` package and a .mmdb file.
    Falls back gracefully — returns ip_df unchanged if anything is missing.
    """
    try:
        import geoip2.database
    except ImportError:
        return ip_df

    if not db_path or not Path(db_path).is_file():
        return ip_df

    countries, states, cities = [], [], []
    reader = geoip2.database.Reader(str(db_path))
    try:
        for ip in ip_df["ipAddress"]:
            try:
                resp = reader.city(ip)
                countries.append(resp.country.name or "")
                states.append(resp.subdivisions.most_specific.name if resp.subdivisions else "")
                cities.append(resp.city.name or "")
            except Exception:
                countries.append("")
                states.append("")
                cities.append("")
    finally:
        reader.close()

    ip_df = ip_df.copy()
    ip_df.insert(ip_df.columns.get_loc("ipAddress") + 1, "country", countries)
    ip_df.insert(ip_df.columns.get_loc("country") + 1, "state", states)
    ip_df.insert(ip_df.columns.get_loc("state") + 1, "city", cities)
    return ip_df


# ── Main IP check entry point ────────────────────────────────────────

def run_ip_check(
    df: pd.DataFrame,
    username: str,
    api_key: str,
    geoip_db_path: str = None,
    progress_callback=None,
) -> pd.DataFrame:
    """Check unique IPs via Scamalytics and optionally enrich with GeoLite2.

    Returns a DataFrame keyed on ipAddress with risk + geo columns.
    """
    df = ensure_columns(df, ["ipAddress"])
    unique_ips = df["ipAddress"].dropna().astype(str).unique()
    unique_ips = [ip for ip in unique_ips if ip and ip != "<NA>"]

    rows = []
    for i, ip in enumerate(unique_ips):
        res = _scam_fetch(ip, username, api_key)
        if res["ok"]:
            rows.append(_curate_ip(ip, res["json"]))
        else:
            row = dict(_EMPTY_IP_ROW)
            row["ip"] = ip
            row["error"] = res.get("error", "")
            rows.append(row)
        time.sleep(random.uniform(*_SCAM_DELAY))
        if progress_callback:
            progress_callback((i + 1) / len(unique_ips))

    ip_df = pd.DataFrame(rows)
    ip_df = ip_df.rename(columns={"ip": "ipAddress"})

    # Enrich with GeoLite2 geolocation if database is available
    if geoip_db_path:
        ip_df = _geoip_enrich(ip_df, geoip_db_path)

    return ip_df


def merge_ip_results(df: pd.DataFrame, ip_df: pd.DataFrame) -> pd.DataFrame:
    """Merge selected IP risk columns into the main DataFrame and add
    duplicate-IP flag and IP total-flags column.

    Columns added to df:
      ip_risk, ip_is_datacenter, ip_is_blacklisted, ip_is_bot,
      ip_duplicate, ip_total_flags
    """
    # Select only the columns we want in the main data
    merge_cols = ["ipAddress", "risk", "is_datacenter", "is_blacklisted", "is_bot"]
    merge_cols = [c for c in merge_cols if c in ip_df.columns]
    merged = df.merge(
        ip_df[merge_cols],
        on="ipAddress", how="left", suffixes=("", "_ip"),
    )

    # Rename for clarity
    for old, new in {"risk": "ip_risk", "is_datacenter": "ip_is_datacenter",
                     "is_blacklisted": "ip_is_blacklisted", "is_bot": "ip_is_bot"}.items():
        if old in merged.columns:
            merged = merged.rename(columns={old: new})

    # Duplicate IP flag — True if this ipAddress appears for more than one respondent
    ip_counts = merged["ipAddress"].map(merged["ipAddress"].value_counts())
    merged["ip_duplicate"] = (ip_counts > 1).fillna(False)

    # IP total flags — count of boolean IP signals that fired
    ip_flag_cols = ["ip_is_datacenter", "ip_is_blacklisted", "ip_is_bot", "ip_duplicate"]
    ip_flag_cols = [c for c in ip_flag_cols if c in merged.columns]
    merged["ip_total_flags"] = merged[ip_flag_cols].apply(
        lambda row: sum(1 for v in row if v is True or v == 1), axis=1
    )

    return merged


# ──────────────────────────────────────────────────────────────────────
# Excel export
# ──────────────────────────────────────────────────────────────────────

def _clean_for_excel(frame: pd.DataFrame) -> pd.DataFrame:
    """Convert all pandas nullable dtypes (Int64, string, Float64, etc.)
    to plain Python objects so openpyxl never encounters pd.NA.

    Key detail: iloc assignment preserves the original backing array dtype,
    so pd.NA sneaks back in. The fix is to rebuild the DataFrame column by
    column using dict construction, which forces the new object dtype.
    """
    _nullable_dtypes = (
        pd.Int8Dtype,   pd.Int16Dtype,  pd.Int32Dtype,  pd.Int64Dtype,
        pd.UInt8Dtype,  pd.UInt16Dtype, pd.UInt32Dtype, pd.UInt64Dtype,
        pd.Float32Dtype, pd.Float64Dtype,
        pd.BooleanDtype, pd.StringDtype,
    )
    cleaned = {}
    for col in frame.columns:
        s = frame[col]
        if isinstance(s.dtype, _nullable_dtypes):
            cleaned[col] = s.astype(object).where(s.notna(), other=None)
        else:
            cleaned[col] = s
    return pd.DataFrame(cleaned, index=frame.index)


def _write_df_to_sheet(ws, df: pd.DataFrame) -> None:
    """Write a DataFrame to an openpyxl worksheet using itertuples.
    Safer than dataframe_to_rows for DataFrames with nullable dtypes.
    """
    ws.append(list(df.columns))
    for row in df.itertuples(index=False):
        ws.append(list(row))


def _build_rename_map(config: dict) -> dict:
    """Build a rename map from internal flag column names to human-readable
    labels sourced from the config. Applied to Excel output only — the
    internal DataFrame keeps the original names throughout processing.

    Label format:  "var1, var2: description"
    Examples:
      flag_incon_1  → "Current_1, Current_2: Very satisfied but would not recommend"
      flag_SL_X     → "SL – Parent_4_1: Satisfaction per event-school"
      flag_S2       → "S2: Age outlier"
      flag_speeder  → "Speeder"
    """
    rename = {
        "flag_speeder":       "Speeder",
        "flag_lagger":        "Lagger",
        "flag_maxdiff_speed": "MaxDiff Speeder",
    }

    # Inconsistency checks → "if_var, then_var: label"
    for chk in _safe_list(config.get("inconsistency_checks")):
        if chk.get("manual_check", False):
            continue
        cid      = chk.get("id", "")
        label    = chk.get("label", "")
        if_var   = chk.get("if_var", "")
        then_var = chk.get("then_var", "")
        if cid and label:
            prefix = ", ".join(dict.fromkeys(v for v in [if_var, then_var] if v))
            rename[f"flag_{cid}"] = f"{prefix}: {label}" if prefix else label

    # Straightlining grids → "SL – prefix/label: label"
    for grid in _safe_list(config.get("grid_variables")):
        label  = grid.get("label", "")
        prefix = grid.get("prefix", "")
        if label:
            var_id = prefix if prefix else label
            rename[f"flag_SL_{label}"] = f"SL – {var_id}: {label}"

    # Numeric variables → "var: label"
    for entry in _safe_list(config.get("numeric_variables")):
        var   = entry.get("var", "")
        label = entry.get("label", "")
        if var and label:
            rename[f"flag_{var}"] = f"{var}: {label}"

    return rename


def _reorder_columns(df: pd.DataFrame, datamap_order: list = None) -> pd.DataFrame:
    """Reorder DataFrame columns for readability in Excel:
      1. uuid
      2. ipAddress + IP check columns (if present)
      3. Key identifiers (transid)
      4. Summary columns (total_flags, SL_Count, reason, action, flag_remove, flag_review)
      5. Flag columns (flag_* — excluding those already in summary)
      6. All remaining survey data columns — sorted by Datamap order if provided
    """
    all_cols = list(df.columns)

    if "uuid" not in all_cols:
        raise ValueError(
            "Column 'uuid' not found in the DataFrame. "
            f"Available columns start with: {all_cols[:10]}"
        )

    priority    = ["uuid"]

    # If IP check was run, place ipAddress + IP columns right after uuid
    ip_cols = ["ipAddress", "ip_risk", "ip_is_datacenter", "ip_is_blacklisted",
               "ip_is_bot", "ip_duplicate", "ip_total_flags"]
    ip_cols = [c for c in ip_cols if c in all_cols]
    priority += ip_cols

    priority += ["transid"]

    summary     = ["total_flags", "SL_Count", "affected_questions",
                   "reason", "action", "flag_remove", "flag_review"]
    summary_set = set(summary)
    flag_cols   = sorted([
        c for c in all_cols
        if c.startswith("flag_") and c not in summary_set
    ])
    used = set(priority + summary + flag_cols)
    rest = [c for c in all_cols if c not in used]

    if datamap_order:
        dm_index = {var: idx for idx, var in enumerate(datamap_order)}
        fallback = len(datamap_order)
        rest.sort(key=lambda c: _datamap_sort_key(c, dm_index, fallback))

    ordered = (
        [c for c in priority if c in all_cols] +
        [c for c in summary  if c in all_cols] +
        flag_cols +
        rest
    )
    return df[ordered]


def export_to_excel(
    df: pd.DataFrame,
    oe_df_and_labels,
    config: dict,
    survey_id: str,
    output_dir: str = ".",
    ip_df: pd.DataFrame = None,
) -> str:
    """Write Excel workbook and return the filepath string.

    Sheets:
      A1           – Full dataset, sorted by flag_remove desc, total_flags desc.
      OE Review    – Two header rows (var names + question labels) then data.
      Flagged Only – Flagged rows only, same column order as A1.
      IP Check     – (optional) Deduplicated IP risk results from Scamalytics.
    """
    config = _ensure_dict(config)

    # Unpack oe_df and labels (build_oe_dataframe returns a tuple)
    if isinstance(oe_df_and_labels, tuple):
        oe_df, oe_labels = oe_df_and_labels
    else:
        oe_df     = oe_df_and_labels
        oe_labels = {}

    today    = datetime.now().strftime("%Y-%m-%d")
    filename = f"{survey_id}_data_cleaning_{today}.xlsx"
    filepath = str(Path(output_dir) / filename)

    # ── Column order from config (populated by Chat C from Datamap) ───
    column_order = _safe_list(config.get("column_order"))

    # ── Drop Decipher system columns not needed in output ─────────────
    _drop_prefixes = ("voqtable", "vqtable", "pagetime")
    drop_cols = [c for c in df.columns if c.startswith(_drop_prefixes)]
    if drop_cols:
        df = df.drop(columns=drop_cols)

    # ── Sort ──────────────────────────────────────────────────────────
    sort_cols = [c for c in ["flag_remove", "total_flags"] if c in df.columns]
    df_sorted = (
        df.sort_values(sort_cols, ascending=[False] * len(sort_cols)).copy()
        if sort_cols else df.copy()
    )
    flagged = (
        df_sorted[df_sorted["total_flags"] > 0].copy()
        if "total_flags" in df_sorted.columns
        else pd.DataFrame()
    )

    # ── Reorder columns: uuid → summary → flags → survey data ────────
    dm_order = column_order if column_order else None
    df_sorted = _reorder_columns(df_sorted, dm_order)
    flagged   = _reorder_columns(flagged, dm_order) if not flagged.empty else flagged

    # ── Rename flag columns to human-readable labels ──────────────────
    rename_map = _build_rename_map(config)
    df_sorted  = df_sorted.rename(columns=rename_map)
    flagged    = flagged.rename(columns=rename_map) if not flagged.empty else flagged

    # ── Convert pd.NA → None so openpyxl can write every cell ────────
    df_sorted = _clean_for_excel(df_sorted)
    flagged   = _clean_for_excel(flagged)
    oe_df     = _clean_for_excel(oe_df)

    # ── Build OE Review two-row header ───────────────────────────────
    oe_entries         = _safe_list(config.get("oe_variables"))
    var_name_row       = ["uuid"] + [v["var"]                 for v in oe_entries]
    question_label_row = ["ID"]   + [v.get("label", v["var"]) for v in oe_entries]

    # ── Write workbook ────────────────────────────────────────────────
    wb = Workbook()

    # Sheet 1: A1 — full dataset
    ws_a1       = wb.active
    ws_a1.title = "A1"
    _write_df_to_sheet(ws_a1, df_sorted)

    # Sheet 2: OE Review — two-row header then data
    ws_oe = wb.create_sheet("OE Review")
    ws_oe.append(var_name_row)
    ws_oe.append(question_label_row)
    for row in oe_df.itertuples(index=False):
        ws_oe.append(list(row))

    # Sheet 3: Flagged Only
    ws_flag = wb.create_sheet("Flagged Only")
    _write_df_to_sheet(ws_flag, flagged)

    # Sheet 4: IP Check (optional)
    if ip_df is not None and not ip_df.empty:
        ip_clean = _clean_for_excel(ip_df)
        ws_ip = wb.create_sheet("IP Check")
        _write_df_to_sheet(ws_ip, ip_clean)

    wb.save(filepath)
    return filepath
